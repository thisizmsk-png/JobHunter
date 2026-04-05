#!/usr/bin/env python3
"""Export vendor_sites.xlsx → data/vendor_urls.json (one-time setup)."""
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Install openpyxl first: pip3 install openpyxl")
    sys.exit(1)

XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "vendor_sites.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "vendor_urls.json")


def export():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)

    vendors = []
    seen_urls = set()

    # Read from master list sheet
    for sheet_name in wb.sheetnames:
        if "Master" in sheet_name or "URL List" in sheet_name:
            ws = wb[sheet_name]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:  # header
                    continue
                if "URL List" in sheet_name:
                    url = row[0]
                    if url and url != "url" and url not in seen_urls:
                        seen_urls.add(url)
                        vendors.append({"url": url, "name": "", "url_type": "confirmed"})
                else:
                    idx, name, url, url_type = (row + (None, None, None, None))[:4]
                    if url and url not in seen_urls and name:
                        seen_urls.add(url)
                        vtype = "confirmed"
                        if url_type and "Inferred" in str(url_type):
                            vtype = "inferred"
                        elif url_type and "New" in str(url_type):
                            vtype = "new"
                        vendors.append({
                            "name": str(name).strip(),
                            "url": str(url).strip(),
                            "url_type": vtype,
                        })

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(vendors, f, indent=2)

    print(f"Exported {len(vendors)} vendor URLs to {OUTPUT_PATH}")


if __name__ == "__main__":
    export()
